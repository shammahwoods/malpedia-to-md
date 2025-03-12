import requests
import json
import os
import re
import logging
import time
import threading
import queue
from datetime import datetime, timedelta
import openpyxl

# ANSI color codes
ANSI_CYAN = "\033[96m"
ANSI_RESET = "\033[0m"

def extract_platform_from_id(family_id):
    """Extract platform from family ID (e.g., 'win.agent_racoon' -> 'Windows')"""
    platform_map = {
        'win': 'Windows',
        'osx': 'macOS',
        'ios': 'iOS',
        'android': 'Android',
        'elf': 'Linux/Unix Executable',
        'aix': 'AIX',
        'apk': 'Android Package',
        'asp': 'ASP Web',
        'fas': 'Flash ActionScript',
        'jar': 'Java Archive',
        'js': 'JavaScript',
        'jsp': 'Java Server Pages',
        'php': 'PHP Web',
        'pl': 'Perl',
        'ps1': 'PowerShell',
        'py': 'Python',
        'sh': 'Shell Script',
        'symbian': 'Symbian OS',
        'vbs': 'Visual Basic Script'
    }
    
    if '.' in family_id:
        platform_code = family_id.split('.')[0]
        return platform_map.get(platform_code, platform_code.capitalize())
    return "Unknown"

def get_platform_folder_name(platform):
    """Convert platform name to folder name"""
    # Remove any parentheses or slashes for folder names
    return re.sub(r'[/\\()]', ' ', platform).strip().replace('  ', ' ')

def get_platform_tag(platform):
    """Convert platform name to a tag-friendly format"""
    # Remove special characters and spaces, convert to lowercase
    return re.sub(r'[/\\() ]', '_', platform).lower().replace('__', '_')

def sanitize_filename(name):
    """Sanitize a string to be used as a filename"""
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def build_existing_files_map(base_output_dir="Malware"):
    """Scan all directories and build a map of family_id -> filename"""
    existing_files = {}
    
    # Check if base directory exists
    if not os.path.exists(base_output_dir):
        return existing_files
    
    # Scan all platform folders
    for platform_folder in os.listdir(base_output_dir):
        platform_path = os.path.join(base_output_dir, platform_folder)
        if os.path.isdir(platform_path):
            # Scan all files in this platform folder
            for filename in os.listdir(platform_path):
                if filename.endswith('.md'):
                    file_path = os.path.join(platform_path, filename)
                    try:
                        # Try to read the family_id from the file's frontmatter
                        with open(file_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                            match = re.search(r'family_id:\s*([\w.]+)', content)
                            if match:
                                family_id = match.group(1)
                                existing_files[family_id] = filename[:-3]  # Store without .md extension
                    except Exception as e:
                        logging.warning(f"Error reading file {file_path}: {e}")
    
    logging.info(f"Found {len(existing_files)} existing malware family files")
    return existing_files

def build_attribution_map():
    """Build a mapping of alternative group names to their primary names from APT data"""
    attribution_map = {}
    
    # Look for country directories
    country_dirs = ['China', 'Russia', 'North Korea', 'Iran', 'Israel', 'NATO', 'Middle East', 'Others', 'Unknown']
    
    for country in country_dirs:
        if not os.path.exists(country):
            logging.warning(f"Country directory {country} not found, skipping")
            continue
            
        # Scan through all markdown files in the directory
        for filename in os.listdir(country):
            if filename.endswith('.md'):
                file_path = os.path.join(country, filename)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        
                        # Extract the primary name
                        primary_match = re.search(r'\[\[(.*?)\]\]', content)
                        if primary_match:
                            primary_name = primary_match.group(1)
                            
                            # Add the primary name itself
                            attribution_map[primary_name.upper()] = primary_name
                            
                            # Look for alternative names
                            # Pattern matches ## OTHER NAME X followed by content
                            alt_name_matches = re.finditer(r'## OTHER NAME \d+\s*\n(.*?)\s*(?:\n|$)', content)
                            for match in alt_name_matches:
                                alt_name = match.group(1).strip()
                                # Some alt names might have [[]] around them, so remove those
                                alt_name = re.sub(r'\[\[(.*?)\]\]', r'\1', alt_name)
                                attribution_map[alt_name.upper()] = primary_name
                                
                except Exception as e:
                    logging.warning(f"Error processing APT file {file_path}: {e}")
    
    # If the APT data hasn't been created yet, try to load it from the Excel file
    if not attribution_map:
        try:
            # Find the Excel file
            excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and 'APT' in f]
            if excel_files:
                excel_file = excel_files[0]
                logging.info(f"Loading attribution data from Excel file: {excel_file}")
                
                workbook = openpyxl.load_workbook(excel_file)
                
                for sheet_name in country_dirs:
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        for row in sheet.iter_rows(min_row=3, values_only=True):
                            if row[0] and row[0] != '?' and str(row[0]).strip():
                                primary_name = str(row[0]).strip().upper()
                                attribution_map[primary_name] = primary_name
                                
                                # Check for alt names in columns (adjust indices based on your Excel structure)
                                alt_name_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]  # Example indices for OTHER NAME columns
                                for col_idx in alt_name_cols:
                                    if col_idx < len(row) and row[col_idx]:
                                        alt_names = str(row[col_idx]).strip()
                                        # Some cells might contain multiple names separated by commas
                                        for alt_name in alt_names.split(','):
                                            clean_alt_name = alt_name.strip().upper()
                                            if clean_alt_name:
                                                attribution_map[clean_alt_name] = primary_name
        except Exception as e:
            logging.error(f"Error loading attribution data from Excel: {e}")
    
    logging.info(f"Built attribution map with {len(attribution_map)} entries")
    return attribution_map

def resolve_attribution(attribution_list, attribution_map):
    """Resolve attribution names to their primary names using the mapping"""
    resolved_attributions = set()
    
    for attr in attribution_list:
        # Convert to uppercase for case-insensitive matching
        attr_upper = attr.upper()
        if attr_upper in attribution_map:
            resolved_attributions.add(attribution_map[attr_upper])
        else:
            # If not found in the map, keep the original
            resolved_attributions.add(attr)
    
    return list(resolved_attributions)

def generate_markdown_file(family_id, family_info, attribution_map, base_output_dir="Malware"):
    """Generate a single Markdown file for a malware family"""
    name = family_info.get("common_name")
    if not name:
        logging.warning(f"Skipping {family_id} - no common name found")
        return None
        
    platform = extract_platform_from_id(family_id)
    platform_folder = get_platform_folder_name(platform)
    platform_tag = get_platform_tag(platform)
    
    # Create platform-specific folder
    output_dir = os.path.join(base_output_dir, platform_folder)
    os.makedirs(output_dir, exist_ok=True)
    
    filename = sanitize_filename(name)
    file_path = os.path.join(output_dir, f"{filename}.md")
    
    # Double-check if file exists (in case it was created during this run)
    if os.path.exists(file_path):
        logging.info(f"File already exists: {file_path} - skipping")
        return None
    
    alt_names = family_info.get("alt_names", [])
    attribution = family_info.get("attribution", [])
    
    # Resolve attribution to primary names using our mapping
    resolved_attribution = resolve_attribution(attribution, attribution_map)
    
    description = family_info.get("description", "")
    urls = family_info.get("urls", [])
    
    with open(file_path, "w", encoding="utf-8") as f:
        # Add YAML frontmatter with all requested properties
        f.write("---\n")
        f.write(f"type: malware\n")
        f.write(f"platform: {platform}\n")
        f.write(f"family_id: {family_id}\n")
        f.write(f"last_updated: {family_info.get('updated', '')}\n")
        f.write(f"source: Malpedia\n")
        
        # Add attribution as an array in the frontmatter with the resolved names
        if resolved_attribution:
            f.write("attribution:\n")
            for actor in resolved_attribution:
                f.write(f"  - \"[[{actor}]]\"\n")
        
        f.write("---\n\n")
        
        # Main name with double brackets for Obsidian linking
        f.write(f"# [[{name}]]\n\n")
        
        # Alternative names as plain text (no brackets)
        if alt_names:
            f.write("## Alternative Names\n")
            for alt_name in alt_names:
                f.write(f"{alt_name}\n")
            f.write("\n")
        
        # Description
        if description:
            f.write("## Description\n")

            clean_description = description.replace('\r\n', '\n')
            f.write(f"{clean_description}\n\n")
        
        # References
        if urls:
            f.write("## References\n")
            for url in urls:
                f.write(f"- {url}\n")
            f.write("\n")
        
    logging.info(f"Created file: {file_path}")
    return platform_folder

def file_writer_thread(file_queue, stats_dict, attribution_map, stop_event):
    """Thread function to write files from queue"""
    while not (stop_event.is_set() and file_queue.empty()):
        try:
            # Non-blocking queue get with timeout
            family_id, family_info = file_queue.get(timeout=1)
            platform_folder = generate_markdown_file(family_id, family_info, attribution_map)
            
            # Update stats if file was created
            if platform_folder:
                with stats_lock:
                    stats_dict[platform_folder] = stats_dict.get(platform_folder, 0) + 1
            
            file_queue.task_done()
        except queue.Empty:
            continue
        except Exception as e:
            logging.error(f"Error in file writer thread: {str(e)}")
            file_queue.task_done()

def print_progress_update(processed, total, start_time):
    """Print a progress update with estimated time remaining"""
    if processed > 0:
        elapsed = time.time() - start_time
        families_per_sec = processed / elapsed
        remaining = total - processed
        if families_per_sec > 0:
            est_remaining_secs = remaining / families_per_sec
            est_completion = datetime.now() + timedelta(seconds=est_remaining_secs)
            eta = est_completion.strftime("%H:%M:%S")
            
            progress_msg = (
                f"{ANSI_CYAN}PROGRESS: Processed {processed}/{total} families "
                f"({processed/total*100:.1f}%) - "
                f"Remaining: {remaining} families - "
                f"ETA: {eta} "
                f"(~{est_remaining_secs/60:.1f} minutes){ANSI_RESET}"
            )
            logging.info(progress_msg)

def main():
    """Main function to execute the workflow"""
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s',
                        handlers=[
                            logging.FileHandler('malpedia_import.log'),
                            logging.StreamHandler()
                        ])
    
    logging.info("Starting Malpedia data import")
    base_url = "https://malpedia.caad.fkie.fraunhofer.de/api"
    base_output_dir = "Malware"
    
    # Create base output directory
    os.makedirs(base_output_dir, exist_ok=True)
    
    # Build attribution mapping from APT data
    logging.info("Building attribution mapping from APT data...")
    attribution_map = build_attribution_map()
    
    # Build map of existing files
    logging.info("Building map of existing malware files...")
    existing_files_map = build_existing_files_map(base_output_dir)
    
    # Set up threading components
    file_queue = queue.Queue(maxsize=100) 
    stop_event = threading.Event()
    global stats_lock
    stats_lock = threading.Lock()
    stats_dict = {}
    
    # Start file writer thread
    writer_thread = threading.Thread(
        target=file_writer_thread,
        args=(file_queue, stats_dict, attribution_map, stop_event)
    )
    writer_thread.daemon = True
    writer_thread.start()
    
    try:
        # Fetch list of all families
        logging.info("Fetching list of all families...")
        families_list_response = requests.get(f"{base_url}/list/families")
        families_list = families_list_response.json()
        time.sleep(1.1)  # Wait to respect rate limit
        
        # Filter out families that already exist
        families_to_fetch = [family_id for family_id in families_list if family_id not in existing_files_map]
        
        logging.info(f"Found {len(families_list)} total families.")
        logging.info(f"Already have {len(existing_files_map)} families.")
        logging.info(f"Need to fetch {len(families_to_fetch)} new families.")
        
        # Track statistics
        total_fetched = 0
        total_skipped = len(existing_files_map)
        total_to_fetch = len(families_to_fetch)
        start_time = time.time()
        
        # Process each family that doesn't exist yet
        for i, family_id in enumerate(families_to_fetch):
            try:
                # Get family details from API
                family_details_response = requests.get(f"{base_url}/get/family/{family_id}")
                if family_details_response.status_code != 200:
                    logging.warning(f"Failed to get info for {family_id}: {family_details_response.status_code}")
                    continue
                
                family_info = family_details_response.json()
                name = family_info.get("common_name")
                
                if not name:
                    logging.warning(f"Skipping {family_id} - no common name found")
                    continue
                
                logging.info(f"Fetched details for family: {family_id}")
                
                # Add to processing queue
                file_queue.put((family_id, family_info))
                total_fetched += 1
                
                # Print progress update every 10 families
                if total_fetched % 10 == 0:
                    print_progress_update(total_fetched, total_to_fetch, start_time)
                
            except Exception as e:
                logging.error(f"Error processing family {family_id}: {str(e)}")
            
            # Wait between requests to stay under rate limit
            time.sleep(1.1)
        
        # Signal writer thread to finish once all API calls are complete
        stop_event.set()
        
        # Wait for all writing to complete
        file_queue.join()
        writer_thread.join(timeout=30)  # Wait up to 30 seconds for thread to finish
        
        # Log summary
        logging.info("=== Summary ===")
        logging.info(f"Total families: {len(families_list)}")
        logging.info(f"Files skipped (already exist): {total_skipped}")
        logging.info(f"Files created: {total_fetched}")
        
        elapsed_time = time.time() - start_time
        hours, remainder = divmod(elapsed_time, 3600)
        minutes, seconds = divmod(remainder, 60)
        logging.info(f"Total time: {int(hours):02}:{int(minutes):02}:{int(seconds):02}")
        
        # Log counts by platform
        logging.info("=== Files created by platform ===")
        for platform, count in sorted(stats_dict.items()):
            logging.info(f"{platform}: {count} files")
        
    except KeyboardInterrupt:
        logging.info("Operation interrupted by user. Waiting for file writer to complete...")
        stop_event.set()
        writer_thread.join(timeout=30)
        logging.info("Import process terminated")
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}", exc_info=True)
        stop_event.set()
        writer_thread.join(timeout=30)

if __name__ == "__main__":
    main()
